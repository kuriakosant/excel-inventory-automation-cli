'''
    EXCEL INVENTORY AUTOMATION TOOL
               CLI
             V-1.0.0

Author : KYRIAKOS ANTONIADIS
mail : kuriakosant2003@gmail.com    
github : https://github.com/kuriakosant
linkedin : https://www.linkedin.com/in/kyriakos-antoniadis-288444326/

'''

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
from colorama import init, Fore, Style

# Initialize colorama
init(autoreset=True)

# Helper function to check if a file exists
def file_exists(filename):
    return os.path.exists(filename)

# Helper function to get valid input for min/max range
def get_float_input(prompt, default=0.0):
    try:
        value = input(prompt)
        if value == "":
            return default
        return float(value)
    except ValueError:
        print("Invalid input. Please input a valid number.")
        return get_float_input(prompt, default)

# Helper function to get valid percentage input (1-100)
def get_percentage_input(prompt):
    try:
        value = int(input(prompt))
        if 1 <= value <= 100:
            return value
        else:
            print("Please input an integer in the range of 1 to 100.")
            return get_percentage_input(prompt)
    except ValueError:
        print("Please input an integer in the range of 1 to 100.")
        return get_percentage_input(prompt)

# Helper function to get rounding choice
def get_rounding_choice():
    try:
        choice = int(input(
            "Please select how you want to round the value:\n"
            " 1. Down to the nearest integer\n"
            " 2. Up to the nearest integer\n"
            " 3. To 1 decimal point\n"
            " 4. To 2 decimal points:\n "
            " Selection(1-4): "
        ))
        if choice in [1, 2, 3, 4]:
            return choice
        else:
            print("Please select a valid rounding option (1-4).")
            return get_rounding_choice()
    except ValueError:
        print("Please select a valid rounding option (1-4).")
        return get_rounding_choice()

# Function to round a number based on user choice
def round_value(value, rounding_choice):
    if pd.isna(value):  # Check for NaN values
        return value  # Return NaN as is
    if value < 1:
        if rounding_choice == 1:  # Round down
            return value  # Keep the original value if it's less than 1
        elif rounding_choice == 2:  # Round up
            return 1  # Round up to 1 if it's less than 1
        else:  # For decimal rounding (choices 3 and 4)
            return max(value, 0.01)  # Ensure the value is at least 0.01
    else:
        if rounding_choice == 1:
            return max(int(value), 1)  # Round down, but not below 1
        elif rounding_choice == 2:
            return int(value) + 1  # Round up
        elif rounding_choice == 3:
            return max(round(value, 1), 0.1)  # Round to 1 decimal, minimum 0.1
        elif rounding_choice == 4:
            return max(round(value, 2), 0.01)  # Round to 2 decimals, minimum 0.01

# Function to modify values based on the case
def modify_values(df, min_val, max_val, case):
    # Check if 'Αξία' exists in the DataFrame
    if 'Αξία' not in df.columns:
        print("Error: The 'Αξία' column was not found in the file. Please make sure the file has the correct format.")
        return None  # Instead of returning the DataFrame, return None

    # Select rows within the min/max range for column 'Αξία', excluding the last row (assumed to contain the sums)
    selected_rows = df.iloc[:-1][(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val)]
    print(f"Selected {len(selected_rows)} rows in the range Min: {min_val}, Max: {max_val}")
    
    if case == 1:  # Reduce quantity of product
        percentage = get_percentage_input("Please input the percentage to reduce the quantity of the products: ")
        rounding_choice = get_rounding_choice()
        
        # Modify 'Ποσ.1' (Column F) and 'Ποσ.2' (Column H)
        for col in ['Ποσ.1', 'Ποσ.2']:
            df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
            ].apply(lambda x: max(x * (1 - percentage / 100), 0.01))  # Ensure minimum value of 0.01

            # Apply rounding based on user choice
            if rounding_choice in [1, 2]:  # Integer rounding
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: max(round(x), 1) if x >= 1 else x)
            elif rounding_choice == 3:  # 1 decimal point
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: round(x, 1))
            elif rounding_choice == 4:  # 2 decimal points
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: round(x, 2))

    elif case == 2:  # Reduce per-unit value of product
        percentage = get_percentage_input("Please input the percentage to reduce the per-unit value of the products: ")
        rounding_choice = get_rounding_choice()
        
        # Modify 'Τιμή κόστους' (Column I)
        df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'] = df.loc[
            (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'
        ].apply(lambda x: max(x * (1 - percentage / 100), 0.01))  # Ensure minimum value of 0.01

        # Apply rounding based on user choice
        if rounding_choice in [1, 2]:  # Integer rounding
            df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'] = df.loc[
                (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'
            ].apply(lambda x: max(round(x), 1) if x >= 1 else x)
        elif rounding_choice == 3:  # 1 decimal point
            df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'] = df.loc[
                (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'
            ].apply(lambda x: round(x, 1))
        elif rounding_choice == 4:  # 2 decimal points
            df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'] = df.loc[
                (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'
            ].apply(lambda x: round(x, 2))

    elif case == 3:  # Reduce total value of product (both quantity and unit price)
        percentage = get_percentage_input("Please input the percentage to reduce both the quantity and per-unit price of the products: ")
        rounding_choice = get_rounding_choice()

        # First, reduce the columns F (Ποσ.1), H (Ποσ.2), and I (Τιμή κόστους)
        df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Ποσ.1'] = df.loc[
            (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Ποσ.1'
        ].apply(lambda x: max(x * (1 - percentage / 100), 0.01))  # Ensure minimum value of 0.01

        df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Ποσ.2'] = df.loc[
            (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Ποσ.2'
        ].apply(lambda x: max(x * (1 - percentage / 100), 0.01))  # Ensure minimum value of 0.01

        df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'] = df.loc[
            (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Τιμή κόστους'
        ].apply(lambda x: max(x * (1 - percentage / 100), 0.01))  # Ensure minimum value of 0.01

        # Apply rounding based on user choice
        for col in ['Ποσ.1', 'Ποσ.2', 'Τιμή κόστους']:
            if rounding_choice in [1, 2]:  # Integer rounding
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: max(round(x), 1) if x >= 1 else x)
            elif rounding_choice == 3:  # 1 decimal point
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: round(x, 1))
            elif rounding_choice == 4:  # 2 decimal points
                df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col] = df.loc[
                    (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), col
                ].apply(lambda x: round(x, 2))

    # Recalculate 'Αξία' (Column J) as 'Ποσ.1' * 'Τιμή κόστους' (F * I = J)
    df.loc[(df['Αξία'] >= min_val) & (df['Αξία'] <= max_val), 'Αξία'] = df.loc[
        (df['Αξία'] >= min_val) & (df['Αξία'] <= max_val)
    ].apply(lambda row: row['Ποσ.1'] * row['Τιμή κόστους'], axis=1)

    # Recalculate the sums of 'Ποσ.1', 'Ποσ.2', 'Τιμή κόστους', and 'Αξία' excluding the last row
    total_pos1 = df.iloc[:-1]['Ποσ.1'].sum()
    total_pos2 = df.iloc[:-1]['Ποσ.2'].sum()  # Summing column H as requested
    total_cost = df.iloc[:-1]['Τιμή κόστους'].sum()
    total_value = df.iloc[:-1]['Αξία'].sum()

    # Update the last row with the recalculated sums
    df.loc[len(df)-1, 'Ποσ.1'] = total_pos1
    df.loc[len(df)-1, 'Ποσ.2'] = total_pos2  # Writing the sum of column H
    df.loc[len(df)-1, 'Τιμή κόστους'] = total_cost
    df.loc[len(df)-1, 'Αξία'] = total_value
    
    return df

# Function to copy formatting from the original file
def copy_formatting(original_file, new_file):
    original_wb = load_workbook(original_file)
    new_wb = load_workbook(new_file)
    
    original_ws = original_wb.active
    new_ws = new_wb.active

    # Copy font and other styles from the original file to the new file
    for row in original_ws.iter_rows(min_row=1, max_row=original_ws.max_row, min_col=1, max_col=original_ws.max_column):
        for cell in row:
            new_cell = new_ws[cell.coordinate]
            new_cell.font = Font(
                name=cell.font.name, 
                size=cell.font.size, 
                bold=cell.font.bold, 
                italic=cell.font.italic, 
                underline=cell.font.underline, 
                strike=cell.font.strike
            )

    new_wb.save(new_file)

# Function to list all files in the current working directory
def list_files_in_directory():
    print("\nFiles in the current directory:")
    for filename in os.listdir('.'):
        if os.path.isfile(filename):
            size_kb = os.path.getsize(filename) / 1024  # Convert bytes to kilobytes
            print(f"{filename} - Size: {size_kb:.2f} KB")  # Display size in KB with 2 decimal places
    print()  # Add a newline for better readability

    # New CLI for selecting the next action after listing files
    while True:
        action = input(Fore.YELLOW + "Please select your next action:\n 1. Choose a file\n 2. Exit:\n Selection(1-2): " + Style.RESET_ALL)
        if action == '1':
            # Directly call the file selection logic
            while True:
                filename = input("Please place your file in the same directory as this script and give the full name of your file (e.g., filename.xlsx): ")
                if file_exists(filename):
                    size_kb = os.path.getsize(filename) / 1024  # Convert bytes to kilobytes
                    print(f"The file you selected is: {filename} ({size_kb:.2f} KB)")
                    
                    # Read the Excel file into a DataFrame
                    df = pd.read_excel(filename, skiprows=8)  # Skip the first 8 rows as per your requirements

                    next_action = input("Please select your next action:\n 1. Modify values\n 2. Exit:\n Selection(1-2): ")
                    if next_action == '1':
                        print("You will now have to specify what rows you want to modify,\nthis is done by providing a min and a max (a range of product value) from the columns J\n")
                        min_value = get_float_input("Please input your min value (press enter for 0): ", 0.0)
                        max_value = get_float_input("Please input your max value: ")
                        print(f"Min: {min_value}, Max: {max_value}")
                        
                        modification_choice = input(
                            "                                                                                                                                      \n"
                            "*In all cases, the affected rows cells in Column J(Αξια) will be recalculated with F*I=J (Ποσ.1 * Τιμη Κοστους= Αξια)*\n"
                            "*Also , Sums of columns F H I and J will be recalculated in the last row of the document*\n"
                            "                                                                                                                                      \n"
                            "Please select your next action:\n"
                            "                                                                                                                                      \n"
                            " 1. Reduce quantity of products\n"
                            "    (decrease the value of all cells in F(Ποσ.1) and H(Ποσ.2) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " 2. Reduce per-unit value of products\n"
                            "    (decrease the value of all cells in I(Τιμη Κοστους) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " 3. Reduce both quantity and per unit value of products\n"
                            "    (decrease the value of all cells in F(Ποσ.1), H(Ποσ.2) and I(Τιμη κοστους) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " Selection(1-3): "
                        )

                        modified_df = modify_values(df, min_value, max_value, int(modification_choice))
                        
                        if modified_df is not None:  # If the column 'Αξία' exists and modifications are made
                            # Get new file name
                            new_filename = input("Please select a name for your new file (just the file name, not including file type): ")
                            modified_df.to_excel(f"{new_filename}.xlsx", index=False)

                            # Copy the formatting from the original file to the new file
                            copy_formatting(filename, f"{new_filename}.xlsx")
                            
                            print(f"Your new file has been created: {new_filename}.xlsx with original formatting.")
                            
                            # Ask if the user wants to convert another file
                            another_file = input("Do you want to convert another file? (y/n): ").lower()
                            if another_file == 'n':
                                print("Exiting...")
                                return  # Exit the loop and program
                        else:
                            print("The file does not contain the 'Αξία' column. Please try again.")
                            break  # Break out of this loop and start from file selection again.
                    elif next_action == '2':  # Check if the user wants to exit
                        print(Fore.RED + "Exiting..." + Style.RESET_ALL)
                        exit()  # Exit the program
                else:
                    print("Invalid file name, please try again.")
        else:  # Handle exit action
            if action == '2':  # If action 2 (Exit) is selected
                print(Fore.RED + "Exiting..." + Style.RESET_ALL)
                exit()

# CLI for the Excel Modifier Tool
def main():
    print(Fore.CYAN + Style.BRIGHT + "EXCEL INVENTORY FILE MODIFIER (V1) (c) 2024 KYRIAKOS ANTONIADIS \n" + Style.RESET_ALL
          + "                                                      \n")
    print(Fore.YELLOW + "Program rules:\n" + Style.RESET_ALL
          + "               \n") 
    print("IMPORTANT: For the program to work the files you input must have a strict format:\n"
          + Fore.YELLOW + "FILE TYPE: The program only accepts .xlsx files\n" + Style.RESET_ALL
          + "Your list of products must begin on Row 9 (so Row 9 should always contain the 1st product)\n"
          + "The cell in " + Fore.YELLOW + "Column F of row 8" + Style.RESET_ALL + " must contain the value " + Fore.YELLOW + "'Ποσ.1'" + Style.RESET_ALL + "\n" 
          + "The cell in " + Fore.YELLOW + "Column I of row 8" + Style.RESET_ALL + " must contain the value " + Fore.YELLOW + "'Ποσ.2'" + Style.RESET_ALL + "\n" 
          + "The cell in " + Fore.YELLOW + "Column J of row 8" + Style.RESET_ALL + " must contain the value " + Fore.YELLOW + "'Αξια'" + Style.RESET_ALL + "\n" 
          + "The program will only accept files that have this format.\n"
          + Fore.YELLOW + "1. The program will always ignore the first 8 rows" + Style.RESET_ALL + " and start implementing the row range selection from row 9 and on.\n" 
          + "2. The program will recalculate the sum of columns F, I, and J in all iterations\n"
          + "The program will always ignore the content of the files final row\n"
          + "if your final row contains a product that product will not be included in any calculations\n"
          + "(this prevents the sum calculation from including the previous sum value)\n"
          + "                ")
    
    while True:
        action = input(Fore.YELLOW + "Please select your next action:\n 1. Choose a file\n 2. List all files in working directory\n 3. Exit:\n Selection(1-3): " + Style.RESET_ALL)
        if action == '1':
            while True:
                filename = input("Please place your file in the same directory as this script and give the full name of your file (e.g., filename.xlsx): ")
                if file_exists(filename):
                    size_kb = os.path.getsize(filename) / 1024  # Convert bytes to kilobytes
                    print(f"The file you selected is: {filename} ({size_kb:.2f} KB)")
                    
                    # Read the Excel file into a DataFrame
                    df = pd.read_excel(filename, skiprows=8)  # Skip the first 8 rows as per your requirements

                    next_action = input("Please select your next action:\n 1. Modify values\n 2. Exit:\n Selection(1-2): ")
                    if next_action == '1':
                        print("You will now have to specify what rows you want to modify,\nthis is done by providing a min and a max (a range of product value) from the columns J\n")
                        min_value = get_float_input("Please input your min value (press enter for 0): ", 0.0)
                        max_value = get_float_input("Please input your max value: ")
                        print(f"Min: {min_value}, Max: {max_value}")
                        
                        modification_choice = input(
                            "                                                                                                                                      \n"
                            "*In all cases, the affected rows cells in Column J(Αξια) will be recalculated with F*I=J (Ποσ.1 * Τιμη Κοστους= Αξια)*\n"
                            "*Also , Sums of columns F H I and J will be recalculated in the last row of the document*\n"
                            "                                                                                                                                      \n"
                            "Please select your next action:\n"
                            "                                                                                                                                      \n"
                            " 1. Reduce quantity of products\n"
                            "    (decrease the value of all cells in F(Ποσ.1) and H(Ποσ.2) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " 2. Reduce per-unit value of products\n"
                            "    (decrease the value of all cells in I(Τιμη Κοστους) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " 3. Reduce both quantity and per unit value of products\n"
                            "    (decrease the value of all cells in F(Ποσ.1), H(Ποσ.2) and I(Τιμη κοστους) columns in the selected rows by a percentage)\n"
                            "                                                                                                                                      \n"
                            " Selection(1-3): "
                        )

                        modified_df = modify_values(df, min_value, max_value, int(modification_choice))
                        
                        if modified_df is not None:  # If the column 'Αξία' exists and modifications are made
                            # Get new file name
                            new_filename = input("Please select a name for your new file (just the file name, not including file type): ")
                            modified_df.to_excel(f"{new_filename}.xlsx", index=False)

                            # Copy the formatting from the original file to the new file
                            copy_formatting(filename, f"{new_filename}.xlsx")
                            
                            print(f"Your new file has been created: {new_filename}.xlsx with original formatting.")
                            
                            # Ask if the user wants to convert another file
                            another_file = input("Do you want to convert another file? (y/n): ").lower()
                            if another_file == 'n':
                                print(Fore.RED + "Exiting..." + Style.RESET_ALL)
                                return  # Exit the loop and program
                        else:
                            print("The file does not contain the 'Αξία' column. Please try again.")
                            break  # Break out of this loop and start from file selection again.
                    elif next_action == '2':  # Check if the user wants to exit
                        print(Fore.RED + "Exiting..." + Style.RESET_ALL)
                        exit()  # Exit the program
                else:
                    print("Invalid file name, please try again.")
        elif action == '2':
            list_files_in_directory()  # Call the function to list files
        elif action == '3':
            print(Fore.RED + "Exiting..." + Style.RESET_ALL)
            break
        else:
            print(Fore.RED + "Please select a valid action (1, 2, or 3)." + Style.RESET_ALL)

if __name__ == "__main__":
    main()

